from tabulate import tabulate
from utils import tabla_pivote
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Border, Side, Font
import numpy as np
from pandas.tseries.offsets import BDay
import warnings


warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
servicios_completos = ['Base de Datos', 'Catalogación', 'Funcional', 'Otro', 'Software Base', 'SysAdmin']


# Calcular días hábiles del cierre de los casos
def calcular_dias_habiles(fecha_inicio, fecha_fin):
        
        if pd.isnull(fecha_inicio) or pd.isnull(fecha_fin):
            return np.nan
        return np.busday_count(fecha_inicio.date(), fecha_fin.date())


# Ajustar el formato de las columnas
def ajustar_ancho_columnas(writer, sheet_name):
        
        workbook = writer.book
        worksheet = workbook[sheet_name]
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for celda in col:
                try:
                    if len(str(celda.value)) > max_length:
                        max_length = len(str(celda.value))
                except:
                    pass
            ajustar_anchura = (max_length + 2)
            worksheet.column_dimensions[column].width = ajustar_anchura


# Generar gráficas de barras
def generar_grafica_barras(worksheet, data_range, categoria_range, titulo, celda):
            
            chart = BarChart()
            chart.title = titulo
            chart.style = 10
            chart.y_axis.title = 'Casos'
            
            data = Reference(worksheet, min_col=data_range['min_col'], min_row=data_range['min_row'],
                    max_col=data_range['max_col'], max_row=data_range['max_row'])
            
            categories = Reference(worksheet, min_col=categoria_range['min_col'], min_row=categoria_range['min_row'],
                                max_col=categoria_range['max_col'], max_row=categoria_range['max_row'])
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            chart.shape = 4
            chart.width = 30  # Ajustar el ancho de la gráfica
            chart.height = 15  # Ajustar la altura de la gráfica
            worksheet.add_chart(chart, celda)


# Generar reporte en Excel
def generar_reporte_excel(df, ruta_salida, total_casos, casos_bac, casos_cobis, casos_infra):
            

        # Filtra los casos válidos (remueve filas con datos vacíos en las columnas relevantes)
            df = df.dropna(subset=['Fecha/Hora Asignación', 'Fecha\nCierre', 'Asignado a', 'Servicio'])


        # Elimina los espacios en blanco adicionales y caracteres específicos (COLUMNA USUARIO) 
            df['USUARIO'] = df['USUARIO'].str.strip().str.lower()


        # Calcula los días en la columna SLA para cada caso -> ('Casos Mas Demorados')
            df['Fecha/Hora Asignación'] = pd.to_datetime(df['Fecha/Hora Asignación'], errors='coerce')
            df['Fecha\nCierre'] = pd.to_datetime(df['Fecha\nCierre'], errors='coerce')
            df['Dias SLA'] = df.apply(lambda row: calcular_dias_habiles(row['Fecha/Hora Asignación'], row['Fecha\nCierre']), axis=1)


        # Estadísticas generales
            total_casos = df.shape[0]
            casos_bac = df[df['USUARIO'].str.contains('bac', case=False, na=False)].shape[0]
            casos_cobis = df[df['USUARIO'].str.contains('topaz', case=False, na=False)].shape[0]
            casos_infra = df[df['USUARIO'].str.contains('infra', case=False, na=False)].shape[0]
            casos_antes_tres_dias = df[(df['Estado'] == 'Cerrado') & (df['Dias SLA'] <= 3)].shape[0]
            casos_despues_tres_dias = df[(df['Estado'] == 'Cerrado') & (df['Dias SLA'] > 3)].shape[0]


        # Capturar demas casos correspondientes al archivo
            casos_por_servicio = tabla_pivote(df, 'Servicio')
            casos_por_persona = tabla_pivote(df, 'Asignado a')
            casos_por_componente = tabla_pivote(df, 'Componente')
            casos_por_ambiente = tabla_pivote(df, 'Ambiente')


        # Calcular el total de servicios asignados por persona
            casos_por_persona = df.groupby(['Asignado a', 'Servicio']).size().unstack(fill_value=0).reindex(columns=servicios_completos, fill_value=0)
            casos_por_persona = casos_por_persona.astype(int)
            casos_por_persona['Total'] = casos_por_persona.sum(axis=1)
            casos_por_persona = casos_por_persona.reset_index()



        # Obtener los casos que más tiempo han tomado en sacar (Topaz y BAC) -> ('Casos Mas Demorados')
            df['Duracion (minutos)'] = pd.to_numeric(df['Duracion (minutos)'], errors='coerce')
            casos_mas_demorados_topaz = df[df['USUARIO'].str.contains('topaz', case=False, na=False)].nlargest(5, 'Duracion (minutos)')
            casos_mas_demorados_bac = df[df['USUARIO'].str.contains('bac', case=False, na=False)].nlargest(5, 'Duracion (minutos)')

            with pd.ExcelWriter(ruta_salida, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Datos Filtrados', index=False)
                ajustar_ancho_columnas(writer, 'Datos Filtrados')

                resumen = pd.DataFrame({
                    'Descripción': ['Total Casos Atendidos', 'Casos BAC', 'Casos Cobis', 'Casos INFRA', 'Casos Cerrados antes de 3 días', 'Casos Cerrados mayor a 3 días'],
                    'Cantidad': [total_casos, casos_bac, casos_cobis, casos_infra, casos_antes_tres_dias, casos_despues_tres_dias]
                })
                resumen.to_excel(writer, sheet_name='Resumen', index=False)
                ajustar_ancho_columnas(writer, 'Resumen')
                casos_por_servicio.to_excel(writer, sheet_name='Casos por Servicio')
                ajustar_ancho_columnas(writer, 'Casos por Servicio')
                casos_por_persona.to_excel(writer, sheet_name='Casos por Persona',  index=False)
                ajustar_ancho_columnas(writer, 'Casos por Persona')
                casos_por_componente.to_excel(writer, sheet_name='Casos por Componente')
                ajustar_ancho_columnas(writer, 'Casos por Componente')
                casos_por_ambiente.to_excel(writer, sheet_name='Casos por Ambiente')
                ajustar_ancho_columnas(writer, 'Casos por Ambiente')
                

                print(df[df['USUARIO'].str.contains('infra', case=False, na=False)])

            # Añadir gráfica de barras por seccion en el archivo excel
                workbook = writer.book

                hojas_con_graficas = {
                    'Resumen': resumen,
                    'Casos por Servicio': casos_por_servicio,
                    'Casos por Persona': casos_por_persona,
                    'Casos por Componente': casos_por_componente,
                    'Casos por Ambiente': casos_por_ambiente
                    
                }

                for nombre_hoja, data in hojas_con_graficas.items():
                    worksheet = workbook[nombre_hoja]
                    data_range = {
                        'min_col': 2,
                        'min_row': 1,
                        'max_col': 2,
                        'max_row': len(data) + 1
                    }
                    categoria_range = {
                        'min_col': 1,
                        'min_row': 2,
                        'max_col': 1,
                        'max_row': len(data) + 1
                    }
                    if nombre_hoja == 'Casos por Persona':
                        celda = 'I15'  
                    else:
                        celda = 'I6'
                    generar_grafica_barras(worksheet, data_range, categoria_range, f'{nombre_hoja}', celda)



            # Añadir los 5 casos más demorados (Topaz y BAC) -> ('Casos Mas Demorados')

            # Casos Topaz(Cobis)
                startrow_topaz = 0
                casos_mas_demorados_topaz.to_excel(writer, sheet_name='Casos Más Demorados', index=False, startrow=startrow_topaz)
             # Casos BAC
                startrow_bac = startrow_topaz + len(casos_mas_demorados_topaz) + 2
                casos_mas_demorados_bac.to_excel(writer, sheet_name='Casos Más Demorados', index=False, startrow=startrow_bac)
                

            # Añadir las estadísticas de casos cerrados en <= 3 días y > 3 días -> ('Casos Mas Demorados')
                worksheet = writer.sheets['Casos Más Demorados']
                stats_start_row = startrow_bac + len(casos_mas_demorados_bac) + 2
                worksheet.cell(row=stats_start_row, column=1, value="Casos Cerrados antes de 3 días")
                worksheet.cell(row=stats_start_row + 1, column=1, value=casos_antes_tres_dias)
                worksheet.cell(row=stats_start_row, column=2, value="Casos Cerrados mayor a 3 días")
                worksheet.cell(row=stats_start_row + 1, column=2, value=casos_despues_tres_dias)


            # Aplicar bordes y negrita a los títulos de las estadísticas -> ('Casos Mas Demorados')
                thin_border = Border(left=Side(style='thin'), 
                                    right=Side(style='thin'), 
                                    top=Side(style='thin'), 
                                    bottom=Side(style='thin'))
                bold_font = Font(bold=True)
                for col in range(1, 3):
                    cell = worksheet.cell(row=stats_start_row, column=col)
                    cell.border = thin_border
                    cell.font = bold_font

                
                ajustar_ancho_columnas(writer, 'Casos Más Demorados')